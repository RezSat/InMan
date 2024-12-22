import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def adjust_column_widths(ws, df, headers, extra_rows=None):
    """
    Dynamically adjust column widths to ensure all content is visible
    
    Args:
    - ws: Worksheet object
    - df: DataFrame with the main data
    - headers: List of column headers
    - extra_rows: Optional list of additional rows to consider for width calculation
    """
    # Combine data from DataFrame and extra rows if provided
    all_data = df.copy()
    if extra_rows is not None:
        for row in extra_rows:
            all_data.loc[len(all_data)] = row

    # Calculate maximum width for each column
    column_widths = {}
    for col_index, col in enumerate(headers, 1):
        # Check header length
        max_length = len(str(col))
        
        # Check data length for this column
        for row in all_data.itertuples(index=False):
            cell_value = str(row[col_index-1]) if row[col_index-1] is not None else ''
            max_length = max(max_length, len(cell_value))
        
        # Add some padding
        column_widths[col_index] = max_length + 2

    # Apply column widths
    for col, width in column_widths.items():
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

def export_to_excel(data, filename):
    # Prepare a list to hold the rows for the DataFrame
    rows = []
    total_divisions = len(data)
    total_employees = 0
    total_items = 0
    
    # Iterate through the divisions
    for division, employees in data.items():
        division_employee_count = len(employees)
        division_item_count = 0  # Initialize total items for the division
        
        for employee in employees:
            employee_id = employee['Employee ID']
            employee_name = employee['Employee Name']
            items = employee['Items']
            division_item_count += len(items)  # Count items
            
            # If the employee has items, iterate through them
            if items:
                for item in items:
                    rows.append({
                        'Division': division,
                        'Employee Name': employee_name,
                        'Employee ID': employee_id,
                        'Item Name': item['name'],
                        'Unique Key': item['unique_key']
                    })
            else:
                # If no items, still add the employee with empty item fields
                rows.append({
                    'Division': division,
                    'Employee Name': employee_name,
                    'Employee ID': employee_id,
                    'Item Name': '',
                    'Unique Key': ''
                })
        
        # Update total counts
        total_employees += division_employee_count
        total_items += division_item_count
        
        # Add a summary row for the division
        rows.append({
            'Division': division,
            'Employee Name': 'Total Employees:',
            'Employee ID': division_employee_count,
            'Item Name': 'Total Items:',
            'Unique Key': division_item_count
        })
        rows.append({
            'Division': '',
            'Employee Name': '',
            'Employee ID': '',
            'Item Name': '',
            'Unique Key': ''
        })  # Add an empty row for spacing

    # Create a DataFrame from the rows
    df = pd.DataFrame(rows)
    
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Items"
    
    # Write the headers
    headers = ['Division', 'Employee Name', 'Employee ID', 'Item Name', 'Unique Key']
    ws.append(headers)
    
    # Apply header formatting
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Light gray background
    
    # Write the data to the worksheet
    for row in df.itertuples(index=False):
        ws.append(row)
    
    # Merge cells for the Division column
    current_division = None
    start_row = 2  # Start from the second row (first row is headers)
    
    for row in range(2, len(df) + 2):  # Adjust for header row
        if df.iloc[row - 2]['Division'] != current_division:
            if current_division is not None:
                # Merge the previous division cells
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                # Apply formatting to the merged cell
                merged_cell = ws.cell(row=start_row, column=1)
                merged_cell.font = Font(bold=True, size=12)  # Bold and larger font size
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_division = df.iloc[row - 2]['Division']
            start_row = row
    
    # Merge the last division cells
    if current_division is not None:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=len(df) + 1, end_column=1)
        # Apply formatting to the merged cell
        merged_cell = ws.cell(row=start_row, column=1)
        merged_cell.font = Font(bold=True, size=12)  # Bold and larger font size
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Merge cells for the Employee Name and Employee ID columns
    current_employee_id = None
    current_employee_name = None
    start_row = 2  # Reset start_row for employee merging
    
    for row in range(2, len(df) + 2):  # Adjust for header row
        if (df.iloc[row - 2]['Employee ID'] != current_employee_id or
            df.iloc[row - 2]['Employee Name'] != current_employee_name):
            if current_employee_id is not None:
                # Merge the previous employee cells
                ws.merge_cells(start_row=start_row, start_column=2, end_row=row - 1, end_column=2)  # Employee Name
                ws.merge_cells(start_row=start_row, start_column=3, end_row=row - 1, end_column=3)  # Employee ID
            current_employee_id = df.iloc[row - 2]['Employee ID']
            current_employee_name = df.iloc[row - 2]['Employee Name']
            start_row = row
    
    # Merge the last employee cells
    if current_employee_id is not None:
        ws.merge_cells(start_row=start_row, start_column=2, end_row=len(df) + 1, end_column=2)  # Employee Name
        ws.merge_cells(start_row=start_row, start_column=3, end_row=len(df) + 1, end_column=3)  # Employee ID
    
    # Highlight and bold the total columns
    for row in range(len(df) + 1):  # Include header row
        cell_value = ws.cell(row=row + 2, column=2).value  # Adjust for header row
        if cell_value is not None and 'Total Employees:' in cell_value:
            for col in [2, 4]:  # Employee ID and Item Name columns
                cell = ws.cell(row=row + 2, column=col)
                cell. font = Font(bold=True, color="FF0000")  # Bold and red font
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add total summary at the end
    ws.append(['', 'TOTAL DIVISIONS:', total_divisions, 'TOTAL EMPLOYEES COUNT:', total_employees])
    ws.append(['', '', '', 'TOTAL ITEMS COUNT:', total_items])

    # Adjust column widths for better visibility
    adjust_column_widths(ws, df, headers, extra_rows=[
        ['', 'TOTAL DIVISIONS:', total_divisions, 'TOTAL EMPLOYEES COUNT:', total_employees],
        ['', '', '', 'TOTAL ITEMS COUNT:', total_items]
    ])

    # Save the workbook
    wb.save(filename)

# Example usage
data = {
    'Division A': [
        {'Employee ID': 'E001', 'Employee Name': 'Alice', 'Items': [{'name': 'Item 1', 'unique_key': 'K001'}]},
        {'Employee ID': 'E002', 'Employee Name': 'Bob', 'Items': []}
    ],
    'Division B': [
        {'Employee ID': 'E003', 'Employee Name': 'Charlie', 'Items': [{'name': 'Item 2', 'unique_key': 'K002'}]},
        {'Employee ID': 'E004', 'Employee Name': 'David', 'Items': [{'name': 'Item 3', 'unique_key': 'K003'}]}
    ]
}

export_to_excel(data, 'employee_items.xlsx')