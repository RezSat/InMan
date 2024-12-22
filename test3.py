# working out put for divison wise

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from controllers.crud import get_employee_details_with_items

def get_column_width(cell_value):
    """
    Calculate appropriate column width based on content.
    Add padding and account for different character widths.
    
    Args:
        cell_value: The value in the cell
    Returns:
        float: Calculated width for the column
    """
    if cell_value is None:
        return 10  # Default minimum width
    
    # Convert to string and handle numbers
    str_value = str(cell_value)
    
    # Calculate base width (rough approximation)
    # Add extra space for bold text, numbers, special characters
    base_width = len(str_value)
    
    # Add padding for comfortable reading
    padding = 4
    
    # Adjust width based on content type
    if any(c.isupper() for c in str_value):
        base_width *= 1.1  # Upper case letters need more space
    
    if any(not c.isalnum() for c in str_value):
        base_width *= 1.1  # Special characters might need more space
    
    return base_width + padding

def export_to_excel(data, headers, output_file='output.xlsx'):
    """
    Export dictionary data to formatted Excel file.
    
    Args:
        data (dict): Dictionary containing division -> employees -> items hierarchy
        headers (list): List of column headers
        output_file (str): Output Excel file name
    """
    wb = Workbook()
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    total_font = Font(bold=True, underline='single')
    grand_total_font = Font(bold=True, size=11, underline='double')
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Initialize column width trackers
    max_column_widths = {i+1: get_column_width(header) for i, header in enumerate(headers)}
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    current_row = 2
    total_all_employees = 0
    total_all_items = 0
    total_divisions = len(data)
    
    # Process each division
    for division, employees in data.items():
        division_start_row = current_row
        total_division_items = 0
        
        # Update max width for division column
        max_column_widths[1] = max(max_column_widths[1], get_column_width(division))
        
        # Process employees and their items
        for employee in employees:
            items = employee['Items']
            total_division_items += len(items)
            
            # Update max widths for employee columns
            max_column_widths[2] = max(max_column_widths[2], get_column_width(employee['Employee Name']))
            max_column_widths[3] = max(max_column_widths[3], get_column_width(employee['Employee ID']))
            
            employee_start_row = current_row
            
            # Write employee items
            for idx, item in enumerate(items):
                row_data = [
                    division if idx == 0 else '',
                    employee['Employee Name'] if idx == 0 else '',
                    employee['Employee ID'] if idx == 0 else '',
                    item['name'],
                    item['unique_key']
                ]
                
                # Update max widths for item columns
                max_column_widths[4] = max(max_column_widths[4], get_column_width(item['name']))
                max_column_widths[5] = max(max_column_widths[5], get_column_width(item['unique_key']))
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.alignment = center_align
                    cell.border = border
                
                current_row += 1
            
            # Merge employee cells if there are multiple items
            if len(items) > 1:
                for col in [2, 3]:  # Employee Name and Employee ID columns
                    ws.merge_cells(
                        start_row=employee_start_row,
                        start_column=col,
                        end_row=current_row - 1,
                        end_column=col
                    )
                    merged_cell = ws.cell(row=employee_start_row, column=col)
                    merged_cell.alignment = center_align
        
        # Merge cells for division
        if current_row > division_start_row:
            ws.merge_cells(
                start_row=division_start_row,
                start_column=1,
                end_row=current_row - 1,
                end_column=1
            )
            division_cell = ws.cell(row=division_start_row, column=1)
            division_cell.alignment = center_align
            
            # Write division totals
            current_row += 1
            total_text = "Total Employees:"
            items_text = "Total Items:"
            max_column_widths[1] = max(max_column_widths[1], get_column_width(total_text))
            max_column_widths[3] = max(max_column_widths[3], get_column_width(items_text))
            
            ws.cell(row=current_row, column=1, value=total_text).font = total_font
            ws.cell(row=current_row, column=2, value=len(employees)).font = total_font
            ws.cell(row=current_row, column=3, value=items_text).font = total_font
            ws.cell(row=current_row, column=4, value=total_division_items).font = total_font
            
            total_all_employees += len(employees)
            total_all_items += total_division_items
            
            current_row += 2  # Add spacing between divisions
    
    # Write grand totals if there are multiple divisions
    if total_divisions > 1:
        grand_total_texts = [
            "TOTAL DIVISIONS:",
            "TOTAL EMPLOYEES COUNT:",
            "TOTAL ITEMS COUNT:"
        ]
        
        # Update max widths for grand total texts
        for text in grand_total_texts:
            max_column_widths[1] = max(max_column_widths[1], get_column_width(text))
        
        ws.cell(row=current_row, column=1, value=grand_total_texts[0]).font = grand_total_font
        ws.cell(row=current_row, column=2, value=total_divisions).font = grand_total_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=grand_total_texts[1]).font = grand_total_font
        ws.cell(row=current_row, column=2, value=total_all_employees).font = grand_total_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=grand_total_texts[2]).font = grand_total_font
        ws.cell(row=current_row, column=2, value=total_all_items).font = grand_total_font
    
    # Apply the calculated column widths
    for col, width in max_column_widths.items():
        # Ensure minimum width and add some padding for better readability
        final_width = max(10, min(60, width))  # Cap between 10 and 60 characters
        ws.column_dimensions[get_column_letter(col)].width = final_width
    
    wb.save(output_file)
        
employees = get_employee_details_with_items()
division_dict = {}
for employee in employees:
    emp_id = employee['emp_id']
    emp_name = employee['name']
    emp_division = employee['division']
    items = employee.get('items', [])

    # If division is not in the dictionary, add it
    if emp_division not in division_dict:
        division_dict[emp_division] = []

    # Append employee details to the division
    division_dict[emp_division].append({
        'Employee ID': emp_id,
        'Employee Name': emp_name,
        'Items': items  # Store items as a list
    })

data = {
    'Division A': [
        {'Employee ID': 'E001', 'Employee Name': 'Alice', 'Items': [{'name': 'Item 1', 'unique_key': 'K001'}]},
        {'Employee ID': 'E002', 'Employee Name': 'Bob', 'Items': []}
    ],
    'Division B': [
        {'Employee ID': 'E003', 'Employee Name': 'Charlie', 'Items': [{'name': 'Item 2', 'unique_key': 'K002'}]},
        {'Employee ID': 'E004', 'Employee Name': 'David', 'Items': [{'name': 'Item 3', 'unique_key': 'K003'}]}
    ]
}

headers = ['Division', 'Employee Name', 'Employee ID', 'Item Name', 'Unique Key']
export_to_excel(division_dict, headers, output_file='output.xlsx')