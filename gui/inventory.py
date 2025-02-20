import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import COLORS
from controllers import (get_all_employees, get_all_items, get_all_divisions, 
                       get_division_details_with_counts, get_employee_details_with_items)
from controllers.crud import get_all_division_names, get_division
from models.database import SessionLocal
from utils.search import search_divisions, search_employee_items, search_employees, search_items, search_unique_key

class InventoryDisplay:
    def __init__(self, main_frame, inv):
        self.main_frame = main_frame
        self.current_view = None
        self.search_results = {
            "Employees": [],
            "Items": [],
            "Divisions": [],
            "Unique Key": [],
            "Employee Items": []
        }
        # State preservation attributes
        self.preserved_state = {
            "search_type": "Select type",
            "query": "",
            "filters": {}
        }

    def preserve_current_state(self):
            """
            Capture the current search state before recreation
            """
            # Preserve search type
            self.preserved_state["search_type"] = self.search_type.get()
            
            # Preserve search query
            self.preserved_state["query"] = self.search_entry.get()
            
            # Preserve specific filters based on search type
            if self.preserved_state["search_type"] == "Employees":
                if hasattr(self, 'division_filter'):
                    self.preserved_state["filters"]["division"] = self.division_filter.get()
            
            elif self.preserved_state["search_type"] == "Items":
                # Preserve item-specific filters if any
                if hasattr(self, 'status_filter'):
                    self.preserved_state["filters"]["status"] = self.status_filter.get()
                if hasattr(self, 'type_filter'):
                    self.preserved_state["filters"]["type"] = self.type_filter.get()

    def restore_preserved_state(self):
        """
        Restore the preserved search state after recreation
        """
        # Restore search type
        self.search_type.set(self.preserved_state["search_type"])
        
        # Create appropriate filters for the search type
        self.create_advanced_filters(self.preserved_state["search_type"])
        
        # Restore search query
        self.search_entry.delete(0, 'end')
        self.search_entry.insert(0, self.preserved_state["query"])
        
        # Restore specific filters
        if self.preserved_state["search_type"] == "Employees":
            if hasattr(self, 'division_filter') and 'division' in self.preserved_state["filters"]:
                self.division_filter.set(self.preserved_state["filters"]["division"])
        
        elif self.preserved_state["search_type"] == "Items":
            # Restore item-specific filters
            if hasattr(self, 'status_filter') and 'status' in self.preserved_state["filters"]:
                self.status_filter.set(self.preserved_state["filters"]["status"])
            if hasattr(self, 'type_filter') and 'type' in self.preserved_state["filters"]:
                self.type_filter.set(self.preserved_state["filters"]["type"])

    def display_search_results(self):
        """
        Recreate the entire view while preserving search state
        """
        # Preserve current state before clearing
        self.preserve_current_state()
        
        # Clear the main frame
        self.clear_main_frame()
        
        # Recreate the basic structure
        self.create_header()
        self.create_search_panel()
        self.create_results_view()
        
        # Restore the preserved state
        self.restore_preserved_state()
        

    def create_header(self):
        header_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        # Title
        title = ctk.CTkLabel(
            header_frame,
            text="Inventory Search",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=COLORS["white"]
        )
        title.pack(side="left", padx=20)
        
        # Export button
        export_btn = ctk.CTkButton(
            header_frame,
            text="Export Results",
            command=self.export_to_excel,
            fg_color=COLORS["green"],
            hover_color=COLORS["secondary_bg"],
            width=120,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        export_btn.pack(side="right", padx=10)

    def create_search_panel(self):
        search_panel = ctk.CTkFrame(self.main_frame, fg_color=COLORS["secondary_bg"])
        search_panel.pack(fill="x", padx=20, pady=10)
        
        # Create two rows for search controls
        top_row = ctk.CTkFrame(search_panel, fg_color="transparent")
        top_row.pack(fill="x", padx=10, pady=(10, 5))
        
        bottom_row = ctk.CTkFrame(search_panel, fg_color="transparent")
        bottom_row.pack(fill="x", padx=10, pady=(5, 10))
        
        # Configure grid columns for top row
        top_row.grid_columnconfigure(1, weight=1)  # Search entry will expand
        
        # Search type selector with improved styling
        search_types = ["Select type","Employees","Items", "Divisions", "Unique Key", "Employee Items"]
        self.search_type = ctk.CTkComboBox(
            top_row,
            values=search_types,
            width=150,
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["black"],
            border_color=COLORS["ash"],
            button_color=COLORS["pink"],
            button_hover_color=COLORS["darker_pink"],
            command=self.on_search_type_change
        )
        self.search_type.grid(row=0, column=0, padx=(0, 10), sticky="w")
        
        # Search entry with improved styling and expansion
        self.search_entry = ctk.CTkEntry(
            top_row,
            placeholder_text="Search...",
            height=40,
            font=ctk.CTkFont(size=14),
            fg_color=COLORS["black"],
            border_color=COLORS["ash"]
        )
        self.search_entry.grid(row=0, column=1, padx=10, sticky="ew")
        self.search_entry.bind('<Return>', self.perform_search)
        
        # Search button with improved styling
        search_btn = ctk.CTkButton(
            top_row,
            text="Search",
            command=self.perform_search,
            fg_color=COLORS["pink"],
            hover_color=COLORS["darker_pink"],
            width=100,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        search_btn.grid(row=0, column=2, padx=(10, 0), sticky="e")
        
        # Advanced filters frame in bottom row
        self.filters_frame = ctk.CTkFrame(bottom_row, fg_color="transparent")
        self.filters_frame.pack(fill="x", expand=True)
        
        # Configure bottom row for filter expansion
        self.filters_frame.grid_columnconfigure(0, weight=1)
        division_filter = ctk.CTkLabel(
        self.filters_frame,
        text="",
        font=ctk.CTkFont(size=14),
        text_color=COLORS["ash"]
            )
        division_filter.pack(side="left", padx=5)

    def create_results_view(self):
        # Main container with border and rounded corners
        container = ctk.CTkFrame(
            self.main_frame,
            fg_color=COLORS["secondary_bg"],
            corner_radius=15,
            border_width=2,
            border_color=COLORS["white"]
        )
        container.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Scrollable frame for results
        self.results_frame = ctk.CTkScrollableFrame(
            container,
            fg_color="transparent",
            scrollbar_button_color=COLORS["pink"],
            scrollbar_button_hover_color=COLORS["darker_pink"]
        )
        self.results_frame.pack(fill="both", expand=True, padx=5, pady=5)

    def create_advanced_filters(self, search_type):
        # Clear existing filters
        for widget in self.filters_frame.winfo_children():
            widget.destroy()
            
        if search_type == "Employees":
            # Division filter
            divisions = ["All Divisions"] + get_all_division_names()
            self.division_filter = ctk.CTkComboBox(
                self.filters_frame,
                values=divisions,
                width=150
            )
            self.division_filter.pack(side="left", padx=5)

        elif search_type == "Employee Items":
            # Attribute name filter
            attr_name_label = ctk.CTkLabel(
                self.filters_frame,
                text="Attribute Name:",
                font=ctk.CTkFont(size=14)
            )
            attr_name_label.pack(side="left", padx=5)
            self.attr_name_filter = ctk.CTkEntry(
                self.filters_frame,
                placeholder_text="Attribute Name",
                width=150,
                height=40,
                font=ctk.CTkFont(size=14),
                fg_color=COLORS["black"],
                border_color=COLORS["ash"]
            )
            self.attr_name_filter.pack(side="left", padx=5)

            # Attribute value filter
            attr_value_label = ctk.CTkLabel(
                self.filters_frame,
                text="Attribute Value:",
                font=ctk.CTkFont(size=14)
            )
            attr_value_label.pack(side="left", padx=5)
            self.attr_value_filter = ctk.CTkEntry(
                self.filters_frame,
                placeholder_text="Attribute Value",
                width=150,
                height=40,
                font=ctk.CTkFont(size=14),
                fg_color=COLORS["black"],
                border_color=COLORS["ash"]
            )
            self.attr_value_filter.pack(side="left", padx=5)

        elif search_type == "Items":
            # Status filter
            self.status_filter = ctk.CTkComboBox(
                self.filters_frame,
                values=["All Status", "Active", "Retired", "Lost"],
                width=150
            )
            #self.status_filter.pack(side="left", padx=5)
            
            # Type filter
            self.type_filter = ctk.CTkComboBox(
                self.filters_frame,
                values=["All Types", "Common", "Individual"],
                width=150
            )
            #self.type_filter.pack(side="left", padx=5)
            
            # Attribute search
            self.attr_name = ctk.CTkEntry(
                self.filters_frame,
                placeholder_text="Attribute Name",
                width=150
            )
            self.attr_value = ctk.CTkEntry(
                self.filters_frame,
                placeholder_text="Attribute Value",
                width=150
            )
            #self.attr_name.pack(side="left", padx=5)
            #self.attr_value.pack(side="left", padx=5)

    def perform_search(self, event=None):
        search_type = self.search_type.get()
        query = self.search_entry.get()
        
        # Clear existing results in the results frame
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        if search_type == "Employees":
            division_filter = getattr(self, 'division_filter', None)
            division_name = division_filter.get() if division_filter and division_filter.get() != "All Divisions" else None
            results = search_employees(query, division_name, items_need=True)
            self.search_results["Employees"] = results
            self.display_search_results()
            self.display_employee_results(results)

        elif search_type == "Employee Items":
            print("HIITTTT THE BUTTON")
            attr_name = self.attr_name_filter.get() if hasattr(self, 'attr_name_filter') else None
            attr_value = self.attr_value_filter.get() if hasattr(self, 'attr_value_filter') else None

            results = search_employee_items(query, attr_name, attr_value)
            self.search_results["Employee Items"] = results
            self.display_employee_item_results(results)

        elif search_type == "Items":
            status_filter = self.status_filter.get() if hasattr(self, 'status_filter') else None
            is_common_filter = self.is_common_filter.get() if hasattr(self, 'is_common_filter') else None
            # Convert is_common filter to boolean
            is_common = None
            if is_common_filter == "Common":
                is_common = 1
            elif is_common_filter == "Not Common":
                is_common = 0
            status = None if status_filter == "All Status" else status_filter
            self.display_search_results()
            results = search_items(
                query, 
                status=status, 
                is_common=is_common
            )
            self.search_results["Items"] = results
            
            self.display_item_results(results)

        elif search_type == "Divisions":
            self.display_search_results()
            results = search_divisions(query)
            self.search_results["Divisions"] = results
            self.display_division_results(results)

        elif search_type == "Unique Key":
            self.display_search_results()
            results = search_unique_key(query)
            self.search_results["Unique Key"] = results
            self.display_unique_key_results(results)

    def export_to_excel(self):
        """
        Export search results to an Excel file with specific attribute formatting
        """
        # Get current search type
        search_type = self.search_type.get()
        
        # Get results for current search type
        results = self.search_results[search_type]
        
        # Check if there are any results
        if not results:
            messagebox.showerror("Export Error", "No results to export.")
            return
        
        # Open file dialog and export
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if file_path:
            try:
                # Create a new workbook and select the active worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
                
                # Styling
                header_font = Font(bold=True, color="000000")  # Black text
                header_fill = PatternFill(
                    start_color="FFF0F5",  # Light Pink (Lavender Blush)
                    end_color="FFF0F5", 
                    fill_type="solid"
                )
                center_alignment = Alignment(
                    horizontal='center', 
                    vertical='center', 
                    wrap_text=True
                )
                
                # Define border style
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                if search_type == "Items":
                    # Write headers
                    headers = [
                        "Name"
                    ]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    
                    # Current row to write data
                    current_row = 2
                    
                    # Process each item
                    for item in results:
                        # Write item name
                        name_cell = ws.cell(row=current_row, column=1, value=item['name'])
                        name_cell.border = thin_border
                        
                        current_row += 1
                    
                elif search_type == "Employees":
                    current_row = 1
                    for emp in results:
                        # Merge and style Employee ID cell
                        emp_id_cell = ws.cell(row=current_row, column=1, value="Employee ID")
                        emp_id_cell.fill = header_fill
                        emp_id_cell.font = header_font
                        emp_id_cell.alignment = center_alignment
                        emp_id_cell.border = thin_border
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1)

                        emp_id_value_cell = ws.cell(row=current_row, column=2, value=emp['emp_id'])
                        emp_id_value_cell.fill = header_fill
                        emp_id_value_cell.alignment = center_alignment
                        emp_id_value_cell.border = thin_border
                        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2)

                        # Merge and style Employee Name cell
                        name_cell = ws.cell(row=current_row, column=3, value="Employee Name")
                        name_cell.fill = header_fill
                        name_cell.font = header_font
                        name_cell.alignment = center_alignment
                        name_cell.border = thin_border
                        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=3)

                        name_value_cell = ws.cell(row=current_row, column=4, value=emp['name'])
                        name_value_cell.fill = header_fill
                        name_value_cell.alignment = center_alignment
                        name_value_cell.border = thin_border
                        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=4)

                        # Merge and style Division cell
                        division_cell = ws.cell(row=current_row, column=5, value="Division")
                        division_cell.fill = header_fill
                        division_cell.font = header_font
                        division_cell.alignment = center_alignment
                        division_cell.border = thin_border
                        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=5)

                        division_value_cell = ws.cell(row=current_row, column=6, value=emp['division'])
                        division_value_cell.fill = header_fill
                        division_value_cell.alignment = center_alignment
                        division_value_cell.border = thin_border
                        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=6)

                        # Move to the next row for items
                        current_row += 2

                        # Add item headers
                        item_headers = ["Item Name", "Unique Key", "Attributes"]
                        for col, header in enumerate(item_headers, 1):
                            cell = ws.cell(row=current_row, column=col)
                            cell.value = header
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            cell.border = thin_border

                        current_row += 1

                        # Add items for the current employee
                        for item in emp['items']:
                            ws.cell(row=current_row, column=1, value=item['name']).border = thin_border
                            ws.cell(row=current_row, column=2, value=item['unique_key']).border = thin_border
                            ws.cell(row=current_row, column=3, value=", ".join([f"{attr['name']}: {attr['value']}" for attr in item['attributes']])).border = thin_border
                            current_row += 1

                        current_row += 2

                elif search_type == "Employee Items":
                    # Write headers
                    headers = [
                        "Emp Id", "Emp Name", "Item Name", "Unique Key", "Attributes"
                    ]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    
                    # Current row to write data
                    current_row = 2

                    # Process each employee item
                    for item in results:
                        # Write employee item details
                        emp_id_cell = ws.cell(row=current_row, column=1, value=item['emp_id'])
                        emp_id_cell.border = thin_border
                        
                        emp_name_cell = ws.cell(row=current_row, column=2, value=item['emp_name'])
                        emp_name_cell.border = thin_border
                        
                        item_name_cell = ws.cell(row=current_row, column=3, value=item['item_name'])
                        item_name_cell.border = thin_border
                        
                        unique_key_cell = ws.cell(row=current_row, column=4, value=item['unique_key'])
                        unique_key_cell.border = thin_border
                        
                        attributes_cell = ws.cell(row=current_row, column=5, value=item['attributes'])
                        attributes_cell.border = thin_border
                        
                        current_row += 1

                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    # Save the workbook
                    wb.save(file_path)
                    messagebox.showinfo("Export Successful", f"Employee Items report saved to:\n{file_path}")

                elif search_type == "Divisions":
                    # Divisions Export
                    headers = [
                        "Division ID", "Name", "Employee Count", "Item Count"
                    ]
                    
                    # Write headers
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    
                    # Current row to write data
                    current_row = 2
                    
                    # Process each division
                    for division in results:
                        # Write division details
                        div_id_cell = ws.cell(row=current_row, column=1, value=division['division_id'])
                        div_id_cell.border = thin_border
                        
                        name_cell = ws.cell(row=current_row, column=2, value=division['name'])
                        name_cell.border = thin_border
                        
                        emp_count_cell = ws.cell(row=current_row, column=3, value=division['employee_count'])
                        emp_count_cell.border = thin_border
                        
                        item_count_cell = ws.cell(row=current_row, column=4, value=division['item_count'])
                        item_count_cell.border = thin_border
                        
                        current_row += 1
                
                elif search_type == "Unique Key":
                    # Unique Key Export
                    headers = [
                        "Employee ID", "Employee Name", "Item Name", "Unique Key"
                    ]
                    
                    # Write headers
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    
                    # Current row to write data
                    current_row = 2
                    
                    # Process each unique key result
                    for result in results:
                        # Write unique key details
                        emp_id_cell = ws.cell(row=current_row, column=1, value=result['emp_id'])
                        emp_id_cell.border = thin_border
                        
                        emp_name_cell = ws.cell(row=current_row, column=2, value=result['employee_name'])
                        emp_name_cell.border = thin_border
                        
                        item_name_cell = ws.cell(row=current_row, column=3, value=result['item_name'])
                        item_name_cell.border = thin_border
                        
                        unique_key_cell = ws.cell(row=current_row, column=4, value=result['unique_key'])
                        unique_key_cell.border = thin_border
                        
                        current_row += 1                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
                
                # Save the workbook
                wb.save(file_path)
                
                messagebox.showinfo("Export Successful", f"Results exported to {file_path}")
            
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export: {str(e)}")
                
    def display(self):
        self.clear_main_frame()
        self.create_header()
        self.create_search_panel()
        self.create_results_view()

    def clear_main_frame(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()

    def on_search_type_change(self, selected_type):
        """Handle changes in the search type selection."""
        self.create_advanced_filters(selected_type)

    def create_headers_for_current_view(self):
        """Create headers for the current view based on the search type."""
        search_type = self.search_type.get()
        # Clear existing headers if any
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        if search_type == "Employees":
            headers = ["Name", "ID", "Division"]
        elif search_type == "Items":
            headers = ["Item Name"]
        elif search_type == "Divisions":
            headers = ["Division Name", "Employee Count"]
        else:
            headers = ["Search Results"]
        
        for header in headers:
            label = ctk.CTkLabel(self.results_frame, text=header)
            label.pack(side="top", padx=5, pady=5)

    def display_employee_results(self, results):
        # Configure grid columns
        self.results_frame.grid_columnconfigure(0, weight=0, minsize=100)  # ID
        self.results_frame.grid_columnconfigure(1, weight=1, minsize=200)  # Name
        self.results_frame.grid_columnconfigure(2, weight=0, minsize=150)  # Division

        # Create headers
        headers = ["Employee ID", "Name", "Division"]
        for col, header in enumerate(headers):
            header_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
            
            ctk.CTkLabel(
                header_frame,
                text=header,
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

        # Add employee rows
        for idx, employee in enumerate(results, 1):
            # ID Cell
            id_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            id_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                id_frame,
                text=employee['emp_id'],
                font=ctk.CTkFont(size=13)
            ).pack(padx=10, pady=8)
            
            # Name Cell
            name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            name_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                name_frame,
                text=employee['name'],
                font=ctk.CTkFont(size=13)
            ).pack(padx=10, pady=8)
            
            # Division Cell
            div_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            div_frame.grid(row=idx, column=2, padx=2, pady=2, sticky="nsew")
            
            div_tag = ctk.CTkFrame(
                div_frame,
                fg_color=COLORS["pink"],
                corner_radius=6
            )
            div_tag.pack(padx=10, pady=8)
            ctk.CTkLabel(
                div_tag,
                text=employee['division'],
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=8, pady=2)

    def display_item_results(self, results):
        # Configure grid columns
        self.results_frame.grid_columnconfigure(0, weight=1, minsize=200)  # Name

        # Create headers
        headers = ["Name"]
        for col, header in enumerate(headers):
            header_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
            
            ctk.CTkLabel(
                header_frame,
                text=header,
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

        # Add item rows
        for idx, item in enumerate(results, 1):
            # Name Cell
            name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            name_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                name_frame,
                text=item['name'],
                font=ctk.CTkFont(size=13),
                wraplength=300
            ).pack(padx=10, pady=8)


    def display_employee_item_results(self, results):
            # Configure grid columns
            self.results_frame.grid_columnconfigure(0, weight=0, minsize=100)  # Emp Id
            self.results_frame.grid_columnconfigure(1, weight=1, minsize=200)  # Emp Name
            self.results_frame.grid_columnconfigure(2, weight=1, minsize=200)  # Item Name
            self.results_frame.grid_columnconfigure(3, weight=1, minsize=200)  # Unique Key
            self.results_frame.grid_columnconfigure(4, weight=1, minsize=300)  # Attributes

            # Create headers
            headers = ["Emp Id", "Emp Name", "Item Name", "Unique Key", "Attributes"]
            for col, header in enumerate(headers):
                header_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
                header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
                
                ctk.CTkLabel(
                    header_frame,
                    text=header,
                    font=ctk.CTkFont(size=14, weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=10, pady=8)

            # Add employee item rows
            for idx, result in enumerate(results, 1):
                # Emp Id Cell
                emp_id_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
                emp_id_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel (emp_id_frame,
                    text=result['emp_id'],
                    font=ctk.CTkFont(size=13),
                    wraplength=300
                ).pack(padx=10, pady=8)

                # Emp Name Cell
                emp_name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
                emp_name_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    emp_name_frame,
                    text=result['emp_name'],
                    font=ctk.CTkFont(size=13),
                    wraplength=300
                ).pack(padx=10, pady=8)

                # Item Name Cell
                item_name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
                item_name_frame.grid(row=idx, column=2, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    item_name_frame,
                    text=result['item_name'],
                    font=ctk.CTkFont(size=13),
                    wraplength=300
                ).pack(padx=10, pady=8)

                # Unique Key Cell
                unique_key_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
                unique_key_frame.grid(row=idx, column=3, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    unique_key_frame,
                    text=result['unique_key'],
                    font=ctk.CTkFont(size=13),
                    wraplength=300
                ).pack(padx=10, pady=8)

                # Attributes Cell
                attributes_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
                attributes_frame.grid(row=idx, column=4, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    attributes_frame,
                    text=result['attributes'],
                    font=ctk.CTkFont(size=13),
                    wraplength=300
                ).pack(padx=10, pady=8)

    def display_division_results(self, results):
        # Configure grid columns
        self.results_frame.grid_columnconfigure(0, weight=1, minsize=200)  # Name
        self.results_frame.grid_columnconfigure(1, weight=0, minsize=150)  # Employee Count
        self.results_frame.grid_columnconfigure(2, weight=0, minsize=150)  # Item Count

        # Create headers
        headers = ["Division Name", "Employee Count", "Item Count"]
        for col, header in enumerate(headers):
            header_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
            
            ctk.CTkLabel(
                header_frame,
                text=header,
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

        # Add division rows
        for idx, division in enumerate(results, 1):
            # Name Cell
            name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            name_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                name_frame,
                text=division['name'],
                font=ctk.CTkFont(size=13)
            ).pack(padx=10, pady=8)
            
            # Employee Count Cell
            employee_count_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            employee_count_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
            
            employee_count_tag = ctk.CTkFrame(
                employee_count_frame,
                fg_color=COLORS["green"],
                corner_radius=6
            )
            employee_count_tag.pack(padx=10, pady=8)
            ctk.CTkLabel(
                employee_count_tag,
                text=str(division['employee_count']),
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=8, pady=2)

            # Item Count Cell
            item_count_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            item_count_frame.grid(row=idx, column=2, padx=2, pady=2, sticky="nsew")
            
            item_count_tag = ctk.CTkFrame(
                item_count_frame,
                fg_color=COLORS["green"],  # Different color for item count
                corner_radius=6
            )
            item_count_tag.pack(padx=10, pady=8)
            ctk.CTkLabel(
                item_count_tag,
                text=str(division['item_count']),
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=8, pady=2)

    def display_combined_results(self, emp_results, item_results, div_results):
        """
        NOT IN USE
        Display search results in separate, independent containers
        
        Args:
            emp_results (list): List of employee search results
            item_results (list): List of item search results
            div_results (list): List of division search results
        """
        # Clear existing results frame
        for widget in self.results_frame.winfo_children():
            widget.destroy()

        # Create a scrollable main container
        main_container = ctk.CTkScrollableFrame(
            self.results_frame, 
            fg_color="transparent",
            scrollbar_button_color=COLORS["pink"],
            scrollbar_button_hover_color=COLORS["darker_pink"]
        )
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # Employees Results Container
        if emp_results:
            # Employees Container
            employees_container = ctk.CTkFrame(
                main_container, 
                fg_color=COLORS["secondary_bg"],
                corner_radius=10
            )
            employees_container.pack(fill="x", pady=10)

            # Employees Header
            emp_header = ctk.CTkFrame(employees_container, fg_color=COLORS["black"])
            emp_header.pack(fill="x", padx=2, pady=2)
            ctk.CTkLabel(
                emp_header,
                text="Employees Search Results",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

            # Employees Results Grid
            emp_results_grid = ctk.CTkFrame(employees_container, fg_color="transparent")
            emp_results_grid.pack(fill="x", padx=5, pady=5)

            # Configure grid columns
            emp_results_grid.grid_columnconfigure(0, weight=0, minsize=100)  # ID
            emp_results_grid.grid_columnconfigure(1, weight=1, minsize=200)  # Name
            emp_results_grid.grid_columnconfigure(2, weight=0, minsize=150)  # Division

            # Headers
            headers = ["Employee ID", "Name", "Division"]
            for col, header in enumerate(headers):
                header_frame = ctk.CTkFrame(emp_results_grid, fg_color=COLORS["black"])
                header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    header_frame,
                    text=header,
                    font=ctk.CTkFont(size=14, weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=10, pady=8)

            # Add employee rows
            for idx, employee in enumerate(emp_results, 1):
                # ID Cell
                id_frame = ctk.CTkFrame(emp_results_grid, fg_color=COLORS["black"])
                id_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    id_frame,
                    text=str(employee['emp_id']),
                    font=ctk.CTkFont(size=13)
                ).pack(padx=10, pady=8)
                
                # Name Cell
                name_frame = ctk.CTkFrame(emp_results_grid, fg_color=COLORS["black"])
                name_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    name_frame,
                    text=employee['name'],
                    font=ctk.CTkFont(size=13)
                ).pack(padx=10, pady=8)
                
                # Division Cell
                div_frame = ctk.CTkFrame(emp_results_grid, fg_color=COLORS["black"])
                div_frame.grid(row=idx, column=2, padx=2, pady=2, sticky="nsew")
                
                div_tag = ctk.CTkFrame(
                    div_frame,
                    fg_color=COLORS["pink"],
                    corner_radius=6
                )
                div_tag.pack(padx=10, pady=8)
                ctk.CTkLabel(
                    div_tag,
                    text=employee['division'],
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=8, pady=2)

        # Items Results Container
        if item_results:
            # Items Container
            items_container = ctk.CTkFrame(
                main_container, 
                fg_color=COLORS["secondary_bg"],
                corner_radius=10
            )
            items_container.pack(fill="x", pady=10)

            # Items Header
            items_header = ctk.CTkFrame(items_container, fg_color=COLORS["black"])
            items_header.pack(fill="x", padx=2, pady=2)
            ctk.CTkLabel(
                items_header,
                text="Items Search Results",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

            # Items Results Grid
            items_results_grid = ctk.CTkFrame(items_container, fg_color="transparent")
            items_results_grid.pack(fill="x", padx=5, pady=5)

            # Configure grid columns
            items_results_grid.grid_columnconfigure(0, weight=1, minsize=200)  # Name
            items_results_grid.grid_columnconfigure(1, weight=0, minsize=100)  # Status
            items_results_grid.grid_columnconfigure(2, weight=0, minsize=300)  # Attributes

            # Headers
            headers = ["Item Name", "Status", "Attributes"]
            for col, header in enumerate(headers):
                header_frame = ctk.CTkFrame(items_results_grid, fg_color=COLORS["black"])
                header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    header_frame,
                    text=header,
                    font=ctk.CTkFont(size=14, weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=10, pady=8)

            # Add item rows
            for idx, item in enumerate(item_results, 1):
                # Name Cell
                name_frame = ctk.CTkFrame(items_results_grid, fg_color=COLORS["black"])
                name_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    name_frame,
                    text=item['name'],
                    font=ctk.CTkFont(size=13),
                    wraplength=300
                ).pack(padx=10, pady=8)
                
                # Status Cell
                status_frame = ctk.CTkFrame(items_results_grid, fg_color=COLORS["black"])
                status_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
                
                status_tag = ctk.CTkFrame(
                    status_frame,
                    fg_color={
                        "active": "#4CAF50",
                        "retired": "#FFA726",
                        "lost": "#EF5350"
                    }.get(item.get('status', '').lower(), COLORS["black"]),
                    corner_radius=6
                )
                status_tag.pack(padx=10, pady=8)
                ctk.CTkLabel(
                    status_tag,
                    text=item.get('status', 'N/A').upper(),
                    font=ctk.CTkFont(size=12 , weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=8, pady=2)
                
                # Attributes Cell
                attrs_frame = ctk.CTkFrame(items_results_grid, fg_color=COLORS["black"])
                attrs_frame.grid(row=idx, column=2, padx=2, pady=2, sticky="nsew")
                
                attrs_container = ctk.CTkFrame(attrs_frame, fg_color="transparent")
                attrs_container.pack(padx=10, pady=8, fill="x")
                
                if 'attributes' in item:
                    for attr in item['attributes']:
                        attr_tag = ctk.CTkFrame(
                            attrs_container,
                            fg_color=COLORS["secondary_bg"],
                            corner_radius=6
                        )
                        attr_tag.pack(side="left", padx=2, pady=2)
                        ctk.CTkLabel(
                            attr_tag,
                            text=f"{attr['name']}: {attr['value']}",
                            font=ctk.CTkFont(size=12),
                            text_color=COLORS["white"]
                        ).pack(padx=6, pady=4)

        # Divisions Results Container
        if div_results:
            # Divisions Container
            divisions_container = ctk.CTkFrame(
                main_container, 
                fg_color=COLORS["secondary_bg"],
                corner_radius=10
            )
            divisions_container.pack(fill="x", pady=10)

            # Divisions Header
            div_header = ctk.CTkFrame(divisions_container, fg_color=COLORS["black"])
            div_header.pack(fill="x", padx=2, pady=2)
            ctk.CTkLabel(
                div_header,
                text="Divisions Search Results",
                font=ctk.CTkFont(size=16, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

            # Divisions Results Grid
            div_results_grid = ctk.CTkFrame(divisions_container, fg_color="transparent")
            div_results_grid.pack(fill="x", padx=5, pady=5)

            # Configure grid columns
            div_results_grid.grid_columnconfigure(0, weight=1, minsize=200)  # Name
            div_results_grid.grid_columnconfigure(1, weight=0, minsize=150)  # Employee Count

            # Headers
            headers = ["Division Name", "Employee Count"]
            for col, header in enumerate(headers):
                header_frame = ctk.CTkFrame(div_results_grid, fg_color=COLORS["black"])
                header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    header_frame,
                    text=header,
                    font=ctk.CTkFont(size=14, weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=10, pady=8)

            # Add division rows
            for idx, division in enumerate(div_results, 1):
                # Name Cell
                name_frame = ctk.CTkFrame(div_results_grid, fg_color=COLORS["black"])
                name_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
                ctk.CTkLabel(
                    name_frame,
                    text=division['name'],
                    font=ctk.CTkFont(size=13)
                ).pack(padx=10, pady=8)
                
                # Employee Count Cell
                count_frame = ctk.CTkFrame(div_results_grid, fg_color=COLORS["black"])
                count_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
                
                count_tag = ctk.CTkFrame(
                    count_frame,
                    fg_color=COLORS["green"],
                    corner_radius=6
                )
                count_tag.pack(padx=10, pady=8)
                ctk.CTkLabel(
                    count_tag,
                    text=str(division['employee_count']),
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color=COLORS["white"]
                ).pack(padx=8, pady=2)

        # Optional: Add a message if no results were found
        if not emp_results and not item_results and not div_results:
            no_results_frame = ctk.CTkFrame(main_container, fg_color=COLORS["black"])
            no_results_frame.pack(fill="x", padx=2, pady=(10, 2))
            ctk.CTkLabel(
                no_results_frame,
                text="No results found.",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

    def display_unique_key_results(self, results):
        # Configure grid columns
        self.results_frame.grid_columnconfigure(0, weight=0, minsize=100)  # Employee ID
        self.results_frame.grid_columnconfigure(1, weight=1, minsize=200)  # Employee Name
        self.results_frame.grid_columnconfigure(2, weight=1, minsize=200)  # Item Name
        self.results_frame.grid_columnconfigure(3, weight=1, minsize=200)  # Unique Key

        # Create headers
        headers = ["Employee ID", "Employee Name", "Item Name", "Unique Key"]
        for col, header in enumerate(headers):
            header_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            header_frame.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
            
            ctk.CTkLabel(
                header_frame,
                text=header,
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)

        # Add rows
        for idx, result in enumerate(results, 1):
            # Employee ID Cell
            emp_id_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            emp_id_frame.grid(row=idx, column=0, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                emp_id_frame,
                text=result['emp_id'],
                font=ctk.CTkFont(size=13)
            ).pack(padx=10, pady=8)
            
            # Employee Name Cell
            emp_name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            emp_name_frame.grid(row=idx, column=1, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                emp_name_frame,
                text=result['employee_name'],
                font=ctk.CTkFont(size=13)
            ).pack(padx=10, pady=8)
            
            # Item Name Cell
            item_name_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            item_name_frame.grid(row=idx, column=2, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(
                item_name_frame,
                text=result['item_name'],
                font=ctk.CTkFont(size=13)
            ).pack(padx=10, pady=8)
            
            # Unique Key Cell
            unique_key_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            unique_key_frame.grid(row=idx, column=3, padx=2, pady=2, sticky="nsew")
            
            unique_key_tag = ctk.CTkFrame(
                unique_key_frame,
                fg_color=COLORS["pink"],
                corner_radius=6
            )
            unique_key_tag.pack(padx=10, pady=8)
            ctk.CTkLabel(
                unique_key_tag,
                text=result['unique_key'],
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=8, pady=2)

        # Optional: No results message
        if not results:
            no_results_frame = ctk.CTkFrame(self.results_frame, fg_color=COLORS["black"])
            no_results_frame.grid(row=1, column=0, columnspan=4, padx=2, pady=(10, 2), sticky="nsew")
            ctk.CTkLabel(
                no_results_frame,
                text="No results found.",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color=COLORS["white"]
            ).pack(padx=10, pady=8)