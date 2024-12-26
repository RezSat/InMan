# utils/summary.py

from controllers.crud import *
from .search  import search_items_by_attribute
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox
from controllers.crud import get_employee_details_with_items

# Get all items assigned to an employee
def get_employee_items(db: Session, emp_id: str):
    return db.query(EmployeeItem).join(Item).filter(EmployeeItem.emp_id == emp_id).all()


# Get all items assigned to employees in a specific division
def get_items_by_division(db: Session, division_name: str):
    division = db.query(Division).filter(Division.name == division_name).first()
    if division:
        return db.query(Item).join(EmployeeItem).join(Employee).filter(Employee.division_id == division.division_id).all()
    return []

# Get all item transfers history
def get_item_transfer_history(db: Session):
    return db.query(ItemTransferHistory).all()

# Generate a report of all items assigned to an employee
def generate_employee_report(db: Session, emp_id: str):
    employee = get_employee(db, emp_id)
    if not employee:
        return "Employee not found."
    
    items = get_employee_items(db, emp_id)
    report = f"Report for Employee: {employee.name} ({emp_id})\n\n"
    
    if not items:
        report += "No items assigned.\n"
    else:
        for emp_item in items:
            report += f"Item Name: {emp_item.item.name}, Unique Key: {emp_item.unique_key}\n"
    
    return report


# Generate a report of all items in the system
def generate_item_report(db: Session):
    items = get_all_items(db)
    report = "System-wide Item Report\n\n"
    
    if not items:
        report += "No items in the system.\n"
    else:
        for item in items:
            # Fetch unique keys from all assignments of this item type
            emp_items = db.query(EmployeeItem).filter(EmployeeItem.item_id == item.item_id).all()
            for emp_item in emp_items:
                report += f"Item Name: {item.name}, Unique Key: {emp_item.unique_key}, Is Common: {'Yes' if item.is_common else 'No'}\n"
    
    return report


# Generate a detailed report of items assigned to an employee, including attributes
def generate_detailed_employee_report(db: Session, emp_id: str):
    employee = get_employee(db, emp_id)
    if not employee:
        return "Employee not found."
    
    items = get_employee_items(db, emp_id)
    report = f"Report for Employee: {employee.name} ({emp_id})\n\n"
    
    if not items:
        report += "No items assigned.\n"
    else:
        for emp_item in items:
            report += f"Item Name: {emp_item.item.name}, Unique Key: {emp_item.unique_key}\n"
            attributes = get_item_attributes(db, emp_item.item_id)
            for attr in attributes:
                report += f"  - {attr.name}: {attr.value}\n"
    
    return report


# Generate a report of items with specific attribute details
def generate_attribute_filtered_item_report(db: Session, name: str, value: str):
    items = search_items_by_attribute(db, name, value)
    report = f"Items Report with Attribute {name} = {value}\n\n"
    
    if not items:
        report += "No items found with the specified attribute.\n"
    else:
        for item in items:
            emp_items = db.query(EmployeeItem).filter(EmployeeItem.item_id == item.item_id).all()
            for emp_item in emp_items:
                report += f"Item Name: {item.name}, Unique Key: {emp_item.unique_key}\n"
                attributes = get_item_attributes(db, emp_item.item_id)
                for attr in attributes:
                    report += f"  - {attr.name}: {attr.value}\n"
                report += "\n"
    
    return report

def get_column_width(cell_value):
    """
    Calculate appropriate column width based on content.
    Add padding and account for different character widths.
    
    Args:
        cell_value: The value in the cell
    Returns:
        float: Calculated width for the column
    """
    if cell_value is None:
        return 10  # Default minimum width
    
    # Convert to string and handle numbers
    str_value = str(cell_value)
    
    # Calculate base width (rough approximation)
    # Add extra space for bold text, numbers, special characters
    base_width = len(str_value)
    
    # Add padding for comfortable reading
    padding = 4
    
    # Adjust width based on content type
    if any(c.isupper() for c in str_value):
        base_width *= 1.1  # Upper case letters need more space
    
    if any(not c.isalnum() for c in str_value):
        base_width *= 1.1  # Special characters might need more space
    
    return base_width + padding


def divison_wise_employee_items_to_excel():
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Division-wise Employee Items Report"
    )
    
    # Check if user cancelled the save dialog
    if not file_path:
        return
    
    try:
        headers = ['Division', 'Employee Name', 'Employee ID', 'Item Name', 'Unique Key | Reference ID']
        employees = get_employee_details_with_items()
        division_dict = {}
        for employee in employees:
            emp_id = employee['emp_id']
            emp_name = employee['name']
            emp_division = employee['division']
            items = employee.get('items', [])

            # If division is not in the dictionary, add it
            if emp_division not in division_dict:
                division_dict[emp_division] = []

            # Append employee details to the division
            division_dict[emp_division].append({
                'Employee ID': emp_id,
                'Employee Name': emp_name,
                'Items': items  # Store items as a list
            })


        wb = Workbook()
        ws = wb.active
        data = division_dict
        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        total_font = Font(bold=True, underline='single')
        grand_total_font = Font(bold=True, size=11, underline='double')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Initialize column width trackers
        max_column_widths = {i+1: get_column_width(header) for i, header in enumerate(headers)}
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        current_row = 2
        total_all_employees = 0
        total_all_items = 0
        total_divisions = len(data)
        
        # Process each division
        for division, employees in data.items():
            division_start_row = current_row
            total_division_items = 0
            
            # Update max width for division column
            max_column_widths[1] = max(max_column_widths[1], get_column_width(division))
            
            # Process employees and their items
            for employee in employees:
                items = employee['Items']
                total_division_items += len(items)
                
                # Update max widths for employee columns
                max_column_widths[2] = max(max_column_widths[2], get_column_width(employee['Employee Name']))
                max_column_widths[3] = max(max_column_widths[3], get_column_width(employee['Employee ID']))
                
                employee_start_row = current_row
                
                # Write employee items
                for idx, item in enumerate(items):
                    row_data = [
                        division if idx == 0 else '',
                        employee['Employee Name'] if idx == 0 else '',
                        employee['Employee ID'] if idx == 0 else '',
                        item['name'],
                        item['unique_key']
                    ]
                    
                    # Update max widths for item columns
                    max_column_widths[4] = max(max_column_widths[4], get_column_width(item['name']))
                    max_column_widths[5] = max(max_column_widths[5], get_column_width(item['unique_key']))
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = value
                        cell.alignment = center_align
                        cell.border = border
                    
                    current_row += 1
                
                # Merge employee cells if there are multiple items
                if len(items) > 1:
                    for col in [2, 3]:  # Employee Name and Employee ID columns
                        ws.merge_cells(
                            start_row=employee_start_row,
                            start_column=col,
                            end_row=current_row - 1,
                            end_column=col
                        )
                        merged_cell = ws.cell(row=employee_start_row, column=col)
                        merged_cell.alignment = center_align
            
            # Merge cells for division
            if current_row > division_start_row:
                ws.merge_cells(
                    start_row=division_start_row,
                    start_column=1,
                    end_row=current_row - 1,
                    end_column=1
                )
                division_cell = ws.cell(row=division_start_row, column=1)
                division_cell.alignment = center_align
                
                # Write division totals
                current_row += 1
                total_text = "Total Employees:"
                items_text = "Total Items:"
                max_column_widths[1] = max(max_column_widths[1], get_column_width(total_text))
                max_column_widths[3] = max(max_column_widths[3], get_column_width(items_text))
                
                ws.cell(row=current_row, column=1, value=total_text).font = total_font
                ws.cell(row=current_row, column=2, value=len(employees)).font = total_font
                ws.cell(row=current_row, column=3, value=items_text).font = total_font
                ws.cell(row=current_row, column=4, value=total_division_items).font = total_font
                
                total_all_employees += len(employees)
                total_all_items += total_division_items
                
                current_row += 2  # Add spacing between divisions
        
        # Write grand totals if there are multiple divisions
        if total_divisions > 1:
            grand_total_texts = [
                "TOTAL DIVISIONS:",
                "TOTAL EMPLOYEES COUNT:",
                "TOTAL ITEMS COUNT:"
            ]
            
            # Update max widths for grand total texts
            for text in grand_total_texts:
                max_column_widths[1] = max(max_column_widths[1], get_column_width(text))
            
            ws.cell(row=current_row, column=1, value=grand_total_texts[0]).font = grand_total_font
            ws.cell(row=current_row, column=2, value=total_divisions).font = grand_total_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value=grand_total_texts[1]).font = grand_total_font
            ws.cell(row=current_row, column=2, value=total_all_employees).font = grand_total_font
            current_row += 1
            
            ws.cell(row=current_row, column=1, value=grand_total_texts[2]).font = grand_total_font
            ws.cell(row=current_row, column=2, value=total_all_items).font = grand_total_font
        
        # Apply the calculated column widths
        for col, width in max_column_widths.items():
            # Ensure minimum width and add some padding for better readability
            final_width = max(10, min(60, width))  # Cap between 10 and 60 characters
            ws.column_dimensions[get_column_letter(col)].width = final_width
        
        wb.save(file_path)
        messagebox.showinfo(
            "Export Successful", 
            f"Division-wise Employee Items report saved to:\n{file_path}"
        )
    except Exception as e:
        messagebox.showerror(
            "Export Error", 
            f"Failed to export report:\n{str(e)}"
        )

def employee_id_name_to_excel():
    """
    Export Employee ID and Name to Excel with user-selected save location
    """
    # Prompt user to choose save location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Employee ID and Name Report"
    )
    
    # Check if user cancelled the save dialog
    if not file_path:
        return
    
    try:
        # Get all employees
        employees = get_all_employees()  # Assuming this function exists
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        
        # Headers
        headers = ['Employee ID', 'Name']
        
        # Styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        total_font = Font(bold=True, underline='single')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Initialize column width trackers
        max_column_widths = {i+1: get_column_width(header) for i, header in enumerate(headers)}
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Current row to write data
        current_row = 2
        total_employees = 0
        
        # Process each employee
        for employee in employees:
            # Get employee details
            emp_id = employee['emp_id']
            emp_name = employee['name']
            
            # Update max widths
            max_column_widths[1] = max(max_column_widths[1], get_column_width(emp_id))
            max_column_widths[2] = max(max_column_widths[2], get_column_width(emp_name))
            
            # Write employee row
            row_data = [emp_id, emp_name]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
            
            current_row += 1
            total_employees += 1
        
        # Write total employees count
        current_row += 1
        ws.cell(row=current_row, column=1, value="TOTAL EMPLOYEES COUNT:").font = total_font
        ws.cell(row=current_row, column=2, value=total_employees).font = total_font
        
        # Apply the calculated column widths
        for col, width in max_column_widths.items():
            # Ensure minimum width and add some padding for better readability
            final_width = max(10, min(60, width))  # Cap between 10 and 60 characters
            ws.column_dimensions[get_column_letter(col)].width = final_width
        
        # Save the workbook
        wb.save(file_path)
        
        # Show success message
        messagebox.showinfo(
            "Export Successful", 
            f"Employee ID and Name report saved to:\n{file_path}"
        )
    
    except Exception as e:
        # Show error message if something goes wrong
        messagebox.showerror(
            "Export Error", 
            f"Failed to export report:\n{str(e)}"
        )

def convert_items_to_dicts(items):
        """
        Convert SQLAlchemy Item objects to dictionaries with detailed attributes
        """
        items_data = []
        for item in items:
            item_dict = {
                "item_id": item.item_id,
                "name": item.name,
                "status": item.status,
                "is_common": item.is_common,
                "attributes": []
            }
            
            # Add attributes
            for attr in item.attributes:
                item_dict["attributes"].append({
                    "name": attr.name,
                    "value": attr.value
                })
            
            items_data.append(item_dict)
        
        return items_data


def items_to_excel():
    """
    Export Items with Name, Is Common, and Attributes to Excel
    """
    # Prompt user to choose save location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Items Report"
    )
    
    # Check if user cancelled the save dialog
    if not file_path:
        return
    
    try:
        # Get all items
        items = convert_items_to_dicts(get_all_items())  # Assuming this function exists
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        
        # Styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = [
            "Name", "Is Common", 
            "Attributes Name", "Attributes Value"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Current row to write data
        current_row = 2
        
        # Process each item
        for item in items:
            # Determine number of attribute rows
            num_attributes = len(item.get('attributes', []))
            attribute_rows = max(1, num_attributes)
            
            # Basic item information
            name_cell = ws.cell(row=current_row, column=1, value=item['name'])
            name_cell.border = thin_border
            
            is_common_cell = ws.cell(row=current_row, column=2, value="Yes" if item['is_common'] else "No")
            is_common_cell.border = thin_border
            
            # Merge cells for basic item info
            if attribute_rows > 1:
                for col in [1, 2]:
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=col,
                        end_row=current_row + attribute_rows - 1,
                        end_column=col
                    )
                    # Apply border to merged cell
                    merged_cell = ws.cell(row=current_row, column=col)
                    merged_cell.border = thin_border
            
            # Write attributes
            if item.get('attributes'):
                for idx, attr in enumerate(item['attributes']):
                    row = current_row + idx
                    attr_name_cell = ws.cell(row=row, column=3, value=attr['name'])
                    attr_name_cell.border = thin_border
                    
                    attr_value_cell = ws.cell(row=row, column=4, value=attr['value'])
                    attr_value_cell.border = thin_border
            else:
                # If no attributes, add placeholders
                no_attr_name_cell = ws.cell(row=current_row, column=3, value="-")
                no_attr_name_cell.border = thin_border
                
                no_attr_value_cell = ws.cell(row=current_row, column=4, value="-")
                no_attr_value_cell.border = thin_border
            
            # Move to next set of rows
            current_row += attribute_rows
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        wb.save(file_path)
        
        # Show success message
        messagebox.showinfo(
            "Export Successful", 
            f"Items report saved to:\n{file_path}"
        )
    
    except Exception as e:
        # Show error message if something goes wrong
        messagebox.showerror(
            "Export Error", 
            f"Failed to export report:\n{str(e)}"
        )