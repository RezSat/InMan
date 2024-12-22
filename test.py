# test.py

from models.database import SessionLocal, initialize_database, session_scope
from controllers.crud import *
from controllers.auth import *
from utils.search import *
from utils.summary import *
from gui.ui import *
from models.models import *

def test1():
    # Initialize Database
    #initialize_database()

    # Start a session
    db = SessionLocal()

    # Create Division
    division = create_division(db, name="IT Department")
    print(f"Created Division: {division.name}")

    # Create Employee
    employee = create_employee(db, emp_id="E001", name="Alice", division_id=division.division_id)
    print(f"Created Employee: {employee.name}")

    # Create Item
    item = create_item(db, name="Laptop", unique_key="SN12345", is_common=False)
    print(f"Created Item: {item.name}")

    # Log an Action
    log_entry = log_action(db, action_type="create", details="Created an item", user_id=None)
    print(f"Log Entry: {log_entry.details}")

    # Create User
    user = create_user(db, username="manager", password="password123")
    print(f"User Created: {user.username}")

    # Authenticate User
    auth_user = authenticate_user(db, username="manager", password="password123")
    if auth_user:
        print("Authentication successful")
    else:
        print("Authentication failed")

    # Fetch all employees
    employees = get_all_employees(db)
    for emp in employees:
        print(f"Employee: {emp.name}")

    db.close()

def test2():
    #initialize_database()
    db = SessionLocal()

    # Create test division and employees
    division = create_division(db, name="Finance")
    emp1 = create_employee(db, emp_id="E002", name="Bob", division_id=division.division_id)
    emp2 = create_employee(db, emp_id="E003", name="Charlie", division_id=division.division_id)
    
    # Create test items
    item1 = create_item(db, name="Desktop", unique_key="DT56789", is_common=False)
    item2 = create_item(db, name="Office Chair", unique_key="OC12345", is_common=True)

    # Assign item to employee
    assignment = assign_item_to_employee(db, emp_id=emp1.emp_id, item_id=item1.item_id, is_unique=True)
    print(f"Assigned Item: {assignment.item_id} to Employee: {emp1.emp_id}")

    # Transfer item from emp1 to emp2
    transfer = transfer_item(db, from_emp_id=emp1.emp_id, to_emp_id=emp2.emp_id, item_id=item1.item_id)
    print(f"Transferred Item: {item1.item_id} from Employee: {emp1.emp_id} to Employee: {emp2.emp_id}")

    # Search Employees, Items, Divisions
    print("Search Results for 'Bob':", search_employees(db, "Bob"))
    print("Search Results for 'Desktop':", search_items(db, "Desktop"))
    print("Search Results for 'Finance':", search_divisions(db, "Finance"))

    db.close()

def test3():
    #initialize_database()
    db = SessionLocal()

    # Create test division and employees
    division = create_division(db, name="HR Department")
    emp1 = create_employee(db, emp_id="E004", name="Dave", division_id=division.division_id)
    emp2 = create_employee(db, emp_id="E005", name="Eve", division_id=division.division_id)
    
    # Create test items
    item1 = create_item(db, name="Projector", unique_key="PRJ12345", is_common=True)
    item2 = create_item(db, name="Phone", unique_key="PHN67890", is_common=False)

    # Assign item to employee
    assign_item_to_employee(db, emp_id=emp1.emp_id, item_id=item1.item_id, is_unique=True)

    # Get all items assigned to employee
    items = get_employee_items(db, emp_id=emp1.emp_id)
    print(f"Employee {emp1.emp_id} Items:", [item.name for item in items])

    # Get items by division
    division_items = get_items_by_division(db, division_name="HR Department")
    print(f"Items in HR Department:", [item.name for item in division_items])

    # Update Employee Info
    updated_employee = update_employee(db, emp_id="E004", new_name="David")
    print(f"Updated Employee: {updated_employee.name}")

    # Remove item from employee
    remove_item_from_employee(db, emp_id="E004", item_id=item1.item_id)

    # Get Transfer History
    transfer_history = get_item_transfer_history(db)
    print("Transfer History:", transfer_history)

    # Get Item by Key
    item = get_item_by_key(db, unique_key="PHN67890")
    print(f"Item by Key: {item.name if item else 'Not Found'}")

    # Generate Employee Report
    employee_report = generate_employee_report(db, emp_id="E004")
    print("Employee Report:\n", employee_report)

    # Generate System-wide Item Report
    item_report = generate_item_report(db)
    print("System-wide Item Report:\n", item_report)

    db.close()


def test4():
    initialize_database()
    # updated test case after the new database model.
    db = SessionLocal()

    # Create a test division
    division = create_division(db, name="Engineering")
    print(f"Created Division: {division.name}")

    # Create test employees
    emp1 = create_employee(db, emp_id="E006", name="Frank", division_id=division.division_id)
    emp2 = create_employee(db, emp_id="E007", name="Grace", division_id=division.division_id)
    print(f"Created Employee: {emp1.name}, ID: {emp1.emp_id}")
    print(f"Created Employee: {emp2.name}, ID: {emp2.emp_id}")

    # Create test items with attributes
    item1 = create_item(db, name="Laptop", unique_key="LT12345", is_common=False)
    item2 = create_item(db, name="Router", unique_key="RT54321", is_common=True)
    print(f"Created Item: {item1.name}, Unique Key: {item1.unique_key}")
    print(f"Created Item: {item2.name}, Unique Key: {item2.unique_key}")

    # Add attributes to items
    add_item_attribute(db, item_id=item1.item_id, name="brand", value="Dell")
    add_item_attribute(db, item_id=item1.item_id, name="model", value="Latitude 5400")
    add_item_attribute(db, item_id=item2.item_id, name="brand", value="Cisco")
    add_item_attribute(db, item_id=item2.item_id, name="model", value="RV340")

    # Assign items to employees
    assign_item_to_employee(db, emp_id=emp1.emp_id, item_id=item1.item_id, is_unique=True)
    assign_item_to_employee(db, emp_id=emp2.emp_id, item_id=item2.item_id, is_unique=False)
    print(f"Assigned Item {item1.name} to Employee {emp1.name}")
    print(f"Assigned Item {item2.name} to Employee {emp2.name}")

    # Test searching items by attribute
    laptops = search_items_by_attribute(db, name="brand", value="Dell")
    print(f"Items with brand 'Dell':", [item.name for item in laptops])

    # Test searching employees by item attribute
    employees_with_dell = search_employees_by_item_attribute(db, name="brand", value="Dell")
    print(f"Employees with Dell items:", [emp.name for emp in employees_with_dell])

    # Test searching employees by item name
    employees_with_laptops = search_employees_by_item_name(db, item_name="Laptop")
    print(f"Employees with Laptops:", [emp.name for emp in employees_with_laptops])

    # Generate detailed employee report
    detailed_report = generate_detailed_employee_report(db, emp_id=emp1.emp_id)
    print("Detailed Employee Report:\n", detailed_report)

    # Generate attribute-filtered item report
    attribute_report = generate_attribute_filtered_item_report(db, name="brand", value="Cisco")
    print("Attribute-Filtered Item Report:\n", attribute_report)

    db.close()


def login_user_creation_test():
    #db = SessionLocal()
    create_user(username="manager", password="abc123")
    #db.close()

def get_all_items_():
    db = SessionLocal()
        # Use joinedload to eagerly load the attribute along with the items
    items = db.query(Item).options(joinedload(Item.attributes)).all()
    db.close()
    return items


items = get_all_items_()

def convert_items_to_dicts(items):
    items_data = []
    for item in items:
        item_dict = {
            "item_id": item.item_id,
            "name": item.name,
            "status": item.status,
            "is_common": item.is_common,
            "attributes": []
        }
        
        # Add attributes
        for attr in item.attributes:
            item_dict["attributes"].append({
                "name": attr.name,
                "value": attr.value
            })
        
        items_data.append(item_dict)
    
    return items_data


# Call the functions together within the same session context
def testsomething():
    with session_scope() as db:
        divisions = get_all_divisions()  # This will still be in the session context
        for div in divisions:
            print(div.name)  # Accessing attributes while in the session context



#login_user_creation_test()
"""import pandas as pd

# Create a DataFrame with the test data
data = {
    'EMP_ID': [101, 102, 103, 104, 105, 106],
    'Name': ['Alice Smith', 'Bob Johnson', 'Charlie Lee', 'David Brown', 'Eva White', 'Frank Black'],
    'Division': ['IT', 'HR', '', 'Sales', '', 'Finance']  # Missing division for Eva White
}

df = pd.DataFrame(data)

# Save the DataFrame to an Excel file
file_name = 'employee_import_test.xlsx'
df.to_excel(file_name, index=False)

print(f"Excel file '{file_name}' created successfully.")
"""
#if __name__ == "__main__":
    #app = InventoryApp()
    #app.run()
    # get_all_items()

    #test1()
    #test2()
    #test3()
    #test4()
    #main_loop()

def get_employee_details_with_items():
    """
    Retrieve all employee details along with their associated items
    
    Returns:
        List of dictionaries containing employee and item information
    """
    try:
        with session_scope() as db:
            # Query employees with their items and division in a single query
            employees = (
                db.query(Employee)
                .options(joinedload(Employee.division))  # Load division data
                .all()
            )
            
            # Transform query results into desired format
            employee_details = []
            for emp in employees:
                # Get items through EmployeeItem relationship
                items_query = (
                    db.query(Item)
                    .join(EmployeeItem)
                    .filter(EmployeeItem.emp_id == emp.emp_id)
                    .all()
                )

                # Format items data
                items_data = []
                for item in items_query:
                    # Get all EmployeeItem records for the current item
                    emp_items = (
                        db.query(EmployeeItem)
                        .filter(
                            EmployeeItem.emp_id == emp.emp_id,
                            EmployeeItem.item_id == item.item_id
                        )
                        .all()
                    )
                    
                    # Prepare item data for each EmployeeItem record
                    for emp_item in emp_items:
                        item_data = {
                            "item_id": item.item_id,
                            "name": item.name,
                            "unique_key": emp_item.unique_key,
                            "date_assigned": emp_item.date_assigned.strftime('%Y-%m-%d') if emp_item.date_assigned else None,
                            "is_common": item.is_common
                        }
                        items_data.append(item_data)
                
                # Prepare employee data
                emp_data = {
                    "emp_id": emp.emp_id,
                    "name": emp.name,
                    "division": emp.division.name if emp.division else "Unassigned",
                    "items": items_data
                }
                employee_details.append(emp_data)

  
            db.close()
            
            return employee_details
    
    except Exception as e:
        logger.error(f"Error retrieving employee details: {str(e)}")
        raise



'''
import openpyxl
import os
wb = Workbook()
ws = wb.active
ws.title = "Employee Data"
ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=5)
wb.save(filename="sample.xlsx")
'''

import pandas as pd
from openpyxl import Workbook
from controllers.crud import get_employee_details_with_items  # Adjust the import based on your structure

def export_employee_data_to_excel(file_path: str):
    """
    Export employee data to an Excel sheet, grouping by division and employee.

    Args:
        file_path (str): The path where the Excel file will be saved.
    """
    try:
        # Retrieve employee data with items
        employees = get_employee_details_with_items()

        # Prepare data for export
        division_dict = {}

        # Group employees by division
        for employee in employees:
            emp_id = employee['emp_id']
            emp_name = employee['name']
            emp_division = employee['division']
            items = employee.get('items', [])

            # If division is not in the dictionary, add it
            if emp_division not in division_dict:
                division_dict[emp_division] = []

            # Append employee details to the division
            division_dict[emp_division].append({
                'Employee ID': emp_id,
                'Employee Name': emp_name,
                'Items': items  # Store items as a list
            })
        print(division_dict)
        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Data"

        # Write headers
        headers = ['Division', 'Employee Name', 'Employee ID', 'Item Name', 'Unique Key']
        ws.append(headers)

        # Write data to the worksheet
        current_row = 2  # Start writing data from the second row (after headers)
        for division, employees in division_dict.items():
            for emp in employees:
                emp_id = emp['Employee ID']
                emp_name = emp['Employee Name']
                items = emp['Items']

                if items:
                    # Write the first item with division and employee details
                    ws.append([division, emp_name, emp_id, items[0]['name'], items[0]['unique_key']])
                    # Merge cells for Division, Employee Name, and Employee ID
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1)  # Employee Name
                    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2)  # Employee ID
                    ws.merge_cells(start_row=current_row, start_column=0, end_row=current_row, end_column=0)  # Division

                    # Add remaining items for the same employee
                    for item in items[1:]:
                        ws.append(['', '', '', item['name'], item['unique_key']])
                else:
                    # If no items, still add the employee with empty item fields
                    ws.append([division, emp_name, emp_id, '', ''])
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1)  # Employee Name
                    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=2)  # Employee ID
                    ws.merge_cells(start_row=current_row, start_column=0, end_row=current_row, end_column=0)  # Division

                current_row += 1  # Increment the row for the next entry

        # Save the workbook
        #wb.save(file_path)
        
        print(f"Employee data exported successfully to {file_path}")

    except Exception as e:
        print(f"An error occurred while exporting employee data: {str(e)}")

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def export_to_excel(data, filename):
    # Prepare a list to hold the rows for the DataFrame
    rows = []
    
    # Iterate through the divisions
    for division, employees in data.items():
        for employee in employees:
            employee_id = employee['Employee ID']
            employee_name = employee['Employee Name']
            items = employee['Items']
            
            # If the employee has items, iterate through them
            if items:
                for item in items:
                    rows.append({
                        'Division': division,
                        'Employee Name': employee_name,
                        'Employee ID': employee_id,
                        'Item Name': item['name'],
                        'Unique Key': item['unique_key']
                    })
            else:
                # If no items, still add the employee with empty item fields
                rows.append({
                    'Division': division,
                    'Employee Name': employee_name,
                    'Employee ID': employee_id,
                    'Item Name': '',
                    'Unique Key': ''
                })
    
    # Create a DataFrame from the rows
    df = pd.DataFrame(rows)
    
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Items"
    
    # Write the headers
    headers = ['Division', 'Employee Name', 'Employee ID', 'Item Name', 'Unique Key']
    ws.append(headers)
    
    # Apply header formatting
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Light gray background
    
    # Write the data to the worksheet
    for row in df.itertuples(index=False):
        ws.append(row)
    
    # Merge cells for the Division column
    current_division = None
    start_row = 2  # Start from the second row (first row is headers)
    
    for row in range(2, len(df) + 2):  # Adjust for header row
        if df.iloc[row - 2]['Division'] != current_division:
            if current_division is not None:
                # Merge the previous division cells
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
            current_division = df.iloc[row - 2]['Division']
            start_row = row
    
    # Merge the last division cells
    if current_division is not None:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=len(df) + 1, end_column=1)
    
    # Merge cells for the Employee Name and Employee ID columns
    current_employee_id = None
    current_employee_name = None
    start_row = 2  # Reset start_row for employee merging
    
    for row in range(2, len(df) + 2):  # Adjust for header row
        if (df.iloc[row - 2]['Employee ID'] != current_employee_id or
            df.iloc[row - 2]['Employee Name'] != current_employee_name):
            if current_employee_id is not None:
                # Merge the previous employee cells
                ws.merge_cells(start_row=start_row, start_column=2, end_row=row - 1, end_column=2)  # Employee Name
                ws.merge_cells(start_row=start_row, start_column=3, end_row=row - 1, end_column=3)  # Employee ID
            current_employee_id = df.iloc[row - 2]['Employee ID']
            current_employee_name = df.iloc[row - 2]['Employee Name']
            start_row = row
    
    # Merge the last employee cells
    if current_employee_id is not None:
        ws.merge_cells(start_row=start_row, start_column=2, end_row=len(df) + 1, end_column=2)  # Employee Name
        ws.merge_cells(start_row=start_row, start_column=3, end_row=len(df) + 1, end_column=3)  # Employee ID
    
    # Adjust column widths and apply center alignment
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20  # Set a default width
        for row in range(1, len(df) + 2):  # Include header row
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set row heights for better visibility
    for row in range(1, len(df) + 2):
        ws.row_dimensions[row].height = 20  # Set a default height
    
    # Save the workbook
    wb.save(filename)

# Example usage
data = {
    'offshore': [
        {'Employee ID': '4444t', 'Employee Name': 'chaminda', 'Items': [{'item_id': 2, 'name': 'hello world', 'unique_key': 'gsdgsgsd', 'date_assigned': '2024-12-21', 'is_common': False}]},
        {'Employee ID': 'gretret ', 'Employee Name': 'yehan', 'Items': [{'item_id': 1, 'name': 'desk2', 'unique_key': '23efsgdsgsg', 'date_assigned': '2024-12-21', 'is_common': True}]},
        {'Employee ID': 'gsgsg', 'Employee Name': 'jenul', 'Items': [{'item_id': 2, 'name': 'hello world', 'unique_key': 'dsgdsgsdg', 'date_assigned': '2024-12-21', 'is_common': False}, {'item_id': 2, 'name': 'hello world', 'unique_key': 'sdgsgdgd', 'date_assigned': '2024-12-21', 'is_common': False}, {'item_id': 1, 'name': 'desk2', 'unique_key': 'dzdgdgdsg', 'date_assigned': '2024-12-21', 'is_common': True}]}
    ],
    'collections department': [
        {'Employee ID': '4526sssdh', 'Employee Name': 'dshss', 'Items': [{'item_id': 3, 'name': 'sfgasgsgsg', 'unique_key': 'sdgdsgdsgsg', 'date_assigned': '2024-12-21', 'is_common': True}, {'item_id': 1, 'name': 'desk2', 'unique_key': 'efewtewtwet', 'date_assigned': '2024-12-21', 'is_common': True}]}
    ]
}

export_to_excel(data, 'employee_items_styled.xlsx')